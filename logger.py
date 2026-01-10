import time
import requests
import sqlite3
import os
from pathlib import Path
from datetime import datetime, timezone
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, PieChart, Reference

# ================= CONFIG =================
API_KEY = os.environ["STEAM_API_KEY"]
STEAM_ID = os.environ["STEAM_ID"]
DISCORD_WEBHOOK = os.environ.get("DISCORD_WEBHOOK_URL")

POLL_INTERVAL = os.environ("POLL_INTERVAL")

ACTIVE_STATES = {1, 2, 5, 6}
IDLE_STATES = {3, 4}

# ================= PATHS =================
DATA_DIR = Path("/app/data")
DATA_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = DATA_DIR / "sessions.db"
EXCEL_FILE = DATA_DIR / "steam_sessions.xlsx"
LAST_UPLOAD_FILE = DATA_DIR / "last_discord_upload.txt"

# ================= DATABASE =================
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS segments (
            steamid TEXT,
            name TEXT,
            start_ts TEXT,
            end_ts TEXT,
            duration REAL,
            segment_type TEXT,
            exported INTEGER DEFAULT 0
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS active (
            steamid TEXT PRIMARY KEY,
            name TEXT,
            start_ts TEXT,
            segment_type TEXT
        )
    """)

    conn.commit()
    conn.close()

# ================= STEAM API =================
def get_tracked_ids():
    r = requests.get(
        "https://api.steampowered.com/ISteamUser/GetFriendList/v1/",
        params={"key": API_KEY, "steamid": STEAM_ID},
        timeout=15
    )
    r.raise_for_status()

    ids = {f["steamid"] for f in r.json()["friendslist"]["friends"]}
    ids.add(STEAM_ID)
    return list(ids)

def get_player_summaries(ids):
    r = requests.get(
        "https://api.steampowered.com/ISteamUser/GetPlayerSummaries/v2/",
        params={"key": API_KEY, "steamids": ",".join(ids)},
        timeout=15
    )
    r.raise_for_status()
    return r.json()["response"]["players"]

# ================= EXCEL =================
def init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sessions"

    ws.append([
        "Name",
        "SteamID",
        "Session Start (UTC)",
        "Session End (UTC)",
        "Playing Duration (minutes)"
    ])

    wb.create_sheet("Summary")
    wb.create_sheet("DailyStats")
    wb.create_sheet("HourHeatmap")
    wb.create_sheet("Charts")

    wb.save(EXCEL_FILE)

def ensure_excel():
    if not EXCEL_FILE.exists():
        init_excel()

def append_sessions_batch(rows):
    """
    Batch write multiple sessions to Excel at once to save I/O overhead.
    rows: List of lists containing session data.
    """
    if not rows:
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Sessions"]
    
    for row in rows:
        ws.append(row)
        
    wb.save(EXCEL_FILE)

# ================= ANALYTICS =================
def rebuild_excel():
    wb = load_workbook(EXCEL_FILE)
    sessions = wb["Sessions"]

    summary = wb["Summary"]
    daily = wb["DailyStats"]
    heat = wb["HourHeatmap"]
    charts = wb["Charts"]

    for ws in (summary, daily, heat, charts):
        ws.delete_rows(1, ws.max_row)

    summary.append([
        "Name",
        "Playing Minutes",
        "Online (No Game) Minutes",
        "Idle Minutes",
        "Playing %",
        "Sessions"
    ])

    daily.append(["Date", "Playing Minutes"])
    heat.append(["Name"] + list(range(24)))

    playing = defaultdict(float)
    online = defaultdict(float)
    idle = defaultdict(float)
    sessions_count = defaultdict(int)
    daily_totals = defaultdict(float)
    hour_map = defaultdict(lambda: [0] * 24)

    # Playing sessions
    for name, sid, start, end, dur in sessions.iter_rows(min_row=2, values_only=True):
        playing[name] += dur
        sessions_count[name] += 1

        dt = datetime.fromisoformat(start)
        daily_totals[dt.date()] += dur
        hour_map[name][dt.hour] += 1

    # Other segments (Online/Idle) from DB
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        SELECT name, segment_type, SUM(duration)
        FROM segments
        WHERE segment_type != 'playing'
        GROUP BY name, segment_type
    """)
    for name, seg, dur in cur.fetchall():
        if seg == "online":
            online[name] += dur
        elif seg == "idle":
            idle[name] += dur
    conn.close()

    for name in playing:
        p = playing[name]
        o = online[name]
        i = idle[name]
        total = p + o + i
        pct = (p / total * 100) if total > 0 else 0

        summary.append([
            name,
            round(p, 1),
            round(o, 1),
            round(i, 1),
            round(pct, 1),
            sessions_count[name]
        ])

    for d in sorted(daily_totals):
        daily.append([str(d), round(daily_totals[d], 1)])

    for name, hours in hour_map.items():
        heat.append([name] + hours)

    charts.append(["Name", "Playing Minutes"])
    for r in summary.iter_rows(min_row=2, values_only=True):
        charts.append([r[0], r[1]])

    # Charts
    pie = PieChart()
    pie.add_data(
        Reference(charts, min_col=2, min_row=1, max_row=charts.max_row),
        titles_from_data=True
    )
    pie.set_categories(
        Reference(charts, min_col=1, min_row=2, max_row=charts.max_row)
    )
    pie.title = "Playing Time Share"
    charts.add_chart(pie, "E2")

    line = LineChart()
    line.add_data(
        Reference(daily, min_col=2, min_row=1, max_row=daily.max_row),
        titles_from_data=True
    )
    line.set_categories(
        Reference(daily, min_col=1, min_row=2, max_row=daily.max_row)
    )
    line.title = "Daily Playing Minutes"
    charts.add_chart(line, "E20")

    wb.save(EXCEL_FILE)

# ================= DISCORD =================
def should_upload_today():
    today = datetime.now(timezone.utc).date()
    if LAST_UPLOAD_FILE.exists():
        try:
            last_date = datetime.fromisoformat(LAST_UPLOAD_FILE.read_text()).date()
            if last_date == today:
                return False
        except ValueError:
            pass # Invalid format, overwrite it
            
    LAST_UPLOAD_FILE.write_text(datetime.now(timezone.utc).isoformat())
    return True

def upload_to_discord():
    if not DISCORD_WEBHOOK or not EXCEL_FILE.exists():
        return

    print("Uploading daily report to Discord...")
    with open(EXCEL_FILE, "rb") as f:
        try:
            requests.post(
                DISCORD_WEBHOOK,
                data={"content": "🎮 **Daily Steam Report** (Activity & Analytics)"},
                files={"file": ("steam_sessions.xlsx", f)},
                timeout=30
            )
        except Exception as e:
            print(f"Failed to upload to Discord: {e}")

# ================= MAIN LOOP =================
def main():
    ensure_excel()
    init_db()

    while True:
        try:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()

            now = datetime.now(timezone.utc).isoformat()
            
            # 1. Fetch current Steam states
            try:
                players = get_player_summaries(get_tracked_ids())
            except Exception as api_error:
                print(f"Steam API Error: {api_error}")
                # Close DB and retry loop after sleep
                conn.close()
                time.sleep(POLL_INTERVAL)
                continue

            current = {}
            for p in players:
                state = p.get("personastate", 0)
                has_game = bool(p.get("gameid"))

                if state in ACTIVE_STATES and has_game:
                    seg = "playing"
                elif state in ACTIVE_STATES:
                    seg = "online"
                elif state in IDLE_STATES:
                    seg = "idle"
                else:
                    continue # Offline or invisible

                current[p["steamid"]] = (p["personaname"], seg)

            # 2. Process active sessions (Close ended ones, update current ones)
            for sid, (name, seg) in current.items():
                cur.execute("SELECT segment_type FROM active WHERE steamid=?", (sid,))
                row = cur.fetchone()

                # If state changed (e.g., from Online -> Playing)
                if not row or row[0] != seg:
                    if row:
                        # Close the old state
                        cur.execute(
                            "SELECT start_ts, segment_type FROM active WHERE steamid=?",
                            (sid,)
                        )
                        start, old = cur.fetchone()
                        dur = (
                            datetime.fromisoformat(now)
                            - datetime.fromisoformat(start)
                        ).total_seconds() / 60
                        
                        cur.execute(
                            "INSERT INTO segments VALUES (?,?,?,?,?, ?,0)",
                            (sid, name, start, now, dur, old)
                        )
                        cur.execute("DELETE FROM active WHERE steamid=?", (sid,))

                    # Start the new state
                    cur.execute(
                        "INSERT OR REPLACE INTO active VALUES (?,?,?,?)",
                        (sid, name, now, seg)
                    )

            # 3. Handle users who went completely Offline
            cur.execute("SELECT steamid, name, start_ts, segment_type FROM active")
            for sid, name, start, seg in cur.fetchall():
                if sid not in current:
                    dur = (
                        datetime.fromisoformat(now)
                        - datetime.fromisoformat(start)
                    ).total_seconds() / 60
                    cur.execute(
                        "INSERT INTO segments VALUES (?,?,?,?,?, ?,0)",
                        (sid, name, start, now, dur, seg)
                    )
                    cur.execute("DELETE FROM active WHERE steamid=?", (sid,))

            # 4. Export finished 'playing' sessions to Excel (Batch Process)
            cur.execute("""
                SELECT rowid, steamid, name, start_ts, end_ts, duration
                FROM segments
                WHERE exported=0 AND segment_type='playing'
            """)
            rows = cur.fetchall()

            if rows:
                print(f"Exporting {len(rows)} new session(s) to Excel.")
                # Format for Excel: Name, ID, Start, End, Duration
                excel_rows = [[r[2], r[1], r[3], r[4], round(r[5], 2)] for r in rows]
                
                # Batch write to Excel
                append_sessions_batch(excel_rows)
                
                # Batch update DB
                row_ids = [(r[0],) for r in rows]
                cur.executemany("UPDATE segments SET exported=1 WHERE rowid=?", row_ids)

                # Rebuild charts/stats
                rebuild_excel()

            conn.commit()
            conn.close()

            # 5. Daily Discord Upload (Decoupled from 'rows' check)
            if should_upload_today():
                upload_to_discord()

        except Exception as e:
            print(f"Critical Loop Error: {e}")
            # Ensure DB is closed if it was left open
            try: conn.close()
            except: pass

        time.sleep(POLL_INTERVAL)

# ================= ENTRY =================
if __name__ == "__main__":
    print("Steam daemon running → playing / online / idle tracked")
    main()